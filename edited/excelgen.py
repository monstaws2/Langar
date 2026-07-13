"""
generate_import.py — LangarMotor product import sheet generator
Scans the current folder for product sub-folders (one per product) and
builds an editable Excel file listing every product, its auto-generated
name/sku, its image filenames, and empty columns for you to fill in.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# We are already sitting inside the folder with all the product folders,
# so we scan "." (meaning "this current folder") instead of "products".
PRODUCTS_DIR = "."
OUTPUT_FILE = "products_import.xlsx"
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp")

# Any folder or file name in here will be SKIPPED — it's not a product.
# Add more names here later if you have other non-product folders.
EXCLUDE_NAMES = {
    "products_import.xlsx",
    "edit.py",
    "logo.jpg",
    "github.com_monstaws2_langar",  # adjust to match your exact folder name
    ".git",
    "__pycache__",
}

HEADERS = [
    "name", "slug", "sku", "price", "category", "brand",
    "compatible_models", "description", "stock_quantity",
    "image_1", "image_2", "image_3", "image_folder",
]


def slug_to_name(slug):
    """Turns 'honda-cg-headlight' into 'Honda Cg Headlight'."""
    return " ".join(word.capitalize() for word in slug.split("-"))


def slug_to_sku(slug):
    """Turns 'honda-cg-headlight' into 'HONDACGHEADLIGHT'."""
    return slug.replace("-", "").upper()


def main():
    if not os.path.isdir(PRODUCTS_DIR):
        print(f"ERROR: Folder not found: {PRODUCTS_DIR}")
        return

    # Get every sub-folder in the current directory, except excluded ones
    product_folders = sorted([
        f for f in os.listdir(PRODUCTS_DIR)
        if os.path.isdir(os.path.join(PRODUCTS_DIR, f))
        and f.lower() not in {n.lower() for n in EXCLUDE_NAMES}
    ])

    if not product_folders:
        print("No product folders found. Check EXCLUDE_NAMES or your folder structure.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    ws.append(HEADERS)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = "A2"

    empty_folders = []

    for slug in product_folders:
        folder_path = os.path.join(PRODUCTS_DIR, slug)

        images = sorted([
            f for f in os.listdir(folder_path)
            if f.lower().endswith(IMAGE_EXTENSIONS)
        ])

        if not images:
            empty_folders.append(slug)

        name = slug_to_name(slug)
        sku = slug_to_sku(slug)

        image_1 = images[0] if len(images) > 0 else ""
        image_2 = images[1] if len(images) > 1 else ""
        image_3 = images[2] if len(images) > 2 else ""
        image_folder = slug  # relative path is just the folder name now

        row = [
            name, slug, sku, "", "", "", "", "", "",
            image_1, image_2, image_3, image_folder,
        ]
        ws.append(row)

    for col_idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 45)

    wb.save(OUTPUT_FILE)

    print("=" * 60)
    print("LangarMotor Product Import Sheet — Summary")
    print("=" * 60)
    print(f"Total product folders found: {len(product_folders)}")
    print(f"Excel file saved as: {OUTPUT_FILE}")

    if empty_folders:
        print(f"\nWARNING: {len(empty_folders)} folder(s) have ZERO images:")
        for f in empty_folders:
            print(f"  - {f}")
    else:
        print("\nAll folders contain at least one image. Good to go!")


if __name__ == "__main__":
    main()